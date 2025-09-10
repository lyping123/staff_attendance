import openpyxl , sqlite3,calendar,math
from datetime import datetime, date
from excel_style import border_alignCenter,BoldFont
from attendance_sycn import sycn_attendance

mydb = sqlite3.connect('attendance.db')

cursor=mydb.cursor()
border, alignment_center= border_alignCenter()
FontBold=BoldFont()


def fetch_staffs(cursor):
    cursor.execute("SELECT * FROM staff_list")
    return cursor.fetchall()

def fetch_attendance(cursor, staff_id, date_today):
    qry = f"""
        SELECT GROUP_CONCAT(time_checkin) as timecheckin 
        FROM staff_attendance 
        WHERE DATE(date_checkin)=DATE('{date_today}') 
        AND staff_id='{staff_id}' 
        AND (time_section='morning' OR time_section='afternoon' OR time_section='other')  
        GROUP BY date_checkin
    """
    cursor.execute(qry)
    return cursor.fetchone()

def calculate_lateness(times):
    """Calculate lateness and attendance duration in minutes"""
    time_objects = [datetime.strptime(time, "%H:%M:%S") for time in times]
    timelateness = 0
    timeatten = 0

    if len(time_objects) > 1:
        for i in range(len(time_objects)//2):
            start_time = max(time_objects[i*2], datetime.strptime("08:00:00", "%H:%M:%S"))
            end_time = min(time_objects[i*2+1], datetime.strptime("17:00:00", "%H:%M:%S"))
            lateness_duration = start_time - datetime.strptime("08:00:00", "%H:%M:%S")
            work_duration = end_time - start_time
            timelateness += lateness_duration.total_seconds() // 60
            timeatten += work_duration.total_seconds() // 60
    else:
        start_time = max(time_objects[0], datetime.strptime("08:00:00", "%H:%M:%S"))
        end_time = min(time_objects[-1], datetime.strptime("17:00:00", "%H:%M:%S"))
        lateness_duration = start_time - datetime.strptime("08:00:00", "%H:%M:%S")
        work_duration = end_time - start_time
        timelateness += lateness_duration.total_seconds() // 60
        timeatten += work_duration.total_seconds() // 60

    return math.floor(timelateness), math.floor(timeatten)

def write_header(sheet):
    sheet.column_dimensions["A"].width = 35
    sheet.column_dimensions["N"].width = 15
    current_year = date.today().year
    sheet["A2"] = "Staff Name"
    sheet.merge_cells(start_row=1, start_column=2, end_row=1, end_column=13)
    sheet.cell(row=1, column=2, value=current_year).alignment =alignment_center
    sheet.cell(row=1, column=2).font = FontBold
    
    for month in range(1, 13):
        sheet.cell(row=2, column=month+1, value=calendar.month_abbr[month])
        sheet.cell(row=2, column=month+1).border=border
        sheet.cell(row=2, column=month+1).alignment = alignment_center
        max_col=month+1
    
    sheet.cell(row=2, column=max_col+1, value="Total lateness").border=border
    sheet.cell(row=2, column=max_col+1).alignment = alignment_center

def export_yearly(cursor):
    try:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        write_header(sheet)

        staffs = fetch_staffs(cursor)
        currect_year = date.today().year
        cal = calendar.Calendar()

        for row, staff in enumerate(staffs, start=3):
            staff_id = staff[3]
            staff_name = staff[1]
            sheet.cell(row=row, column=1, value=staff_name).border=border
            
            

            for month in range(1, 13):
                timelateness = 0
                for week in cal.monthdatescalendar(currect_year, month):
                    for day in week:
                        if day.weekday() in [5, 6] or day.month != month:
                            continue
                        record = fetch_attendance(cursor, staff_id, day)
                        if record and record[0]:
                            times = record[0].split(",")
                            late, _ = calculate_lateness(times)
                            timelateness += late

                sheet.cell(row=row, column=month+1, value=timelateness).border=border
                sheet.cell(row=row, column=month+1).alignment = alignment_center
            total_lateness = sum(sheet.cell(row=row, column=col).value or 0 for col in range(2, 14))
            
            sheet.cell(row=row, column=14, value=total_lateness).border=border
            sheet.cell(row=row, column=14).alignment = alignment_center

        workbook.save(f"Staff attendance yearly summary({currect_year}).xlsx")
        print("Export completed successfully")

    except Exception as e:
        print(f"Error occurred: {e}")

export_yearly(cursor)
sycn_attendance()
