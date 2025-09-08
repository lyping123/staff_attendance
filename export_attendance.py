import sqlite3
import tkinter as tk
import openpyxl
from openpyxl.styles import Border, Side
from openpyxl.styles import PatternFill
from tkinter import ttk,messagebox,filedialog
from datetime import datetime,timedelta,date
import calendar
import math
import requests

mydb = sqlite3.connect('attendance.db')

cursor=mydb.cursor()

def adapt_date(date_obj):
    return date_obj.isoformat()

sqlite3.register_adapter(date, adapt_date)

def export_excel():
    try:
        workbook = openpyxl.Workbook()
        
        query="SELECT * FROM staff_list"
        cursor.execute(query)
        staffs=cursor.fetchall()
        for staff in staffs:
            staff_name, staff_department, staff_id = staff[1], staff[2], staff[3]
            sheet_name=f"{staff_name}"
            sheet=workbook.create_sheet(title=sheet_name)
            cal = calendar.Calendar()
            
            currect_month=datetime.today().month
            currect_year=date.today().year
            month_dates = cal.monthdatescalendar(currect_year, currect_month)
            
            sheet["A1"]="DAY"
            sheet["B1"]="Time in"
            sheet["C1"]="Time out"
            sheet["D1"]="Break time"
            sheet["E1"]="Total working hours"
            sheet["F1"]="Total time off"
            
            sheet.column_dimensions["A"].width=5
            sheet.column_dimensions["B"].width=25
            sheet.column_dimensions["C"].width=25
            sheet.column_dimensions["D"].width=25
            sheet.column_dimensions["E"].width=10
            sheet.column_dimensions["F"].width=10
           
            
            for week in month_dates:
                for day in week:
                    if day.month==currect_month:
                        if day.weekday() in [5, 6]: 
                            pass
                        else:
                            
                            # print(f"{day.day:2}")
                            date_today=day
                            qry=f"SELECT GROUP_CONCAT(time_checkin) as timecheckin FROM staff_attendance where DATE(date_checkin)=DATE('{date_today}') AND staff_id='{staff[3]}' AND (time_section='morning' or time_section='afternoon' or time_section='other')  group by date_checkin"
                            cursor.execute(qry)
                            datecheckin=cursor.fetchone()
                            
                            
                            if datecheckin is not None:
                                qry_break=f"SELECT GROUP_CONCAT(time_checkin) as timecheckin FROM staff_attendance where DATE(date_checkin)=DATE('{date_today}') AND staff_id='{staff[3]}' AND time_section='breaktime'  group by date_checkin"
                                cursor.execute(qry_break)
                                breaktimecheck=cursor.fetchone()
                                if breaktimecheck is not None:
                                    breaktimescan=breaktimecheck[0]
                                    breaktimes=breaktimecheck[0].split(",")
                                    time_breaks = [datetime.strptime(breaktime, "%H:%M:%S") for breaktime in breaktimes]
                                    breaktimelen=len(time_breaks)
                                    if breaktimelen==2:
                                        breaktime=time_breaks[-1]-time_breaks[0]
                                        breaktime_duration=math.floor(breaktime.total_seconds() / 60)
                                else:
                                    breaktimescan=""
                                    breaktime_duration=60
                                
                                checktime=datecheckin[0]
                                times=checktime.split(",")
                                time_objects = [datetime.strptime(time, "%H:%M:%S") for time in times]
                                count_timeoff=len(time_objects)
                                timeatten=0
                                timeins = [times[i] for i in range(len(times)) if i % 2 == 0]
                                timeouts = [times[i] for i in range(len(times)) if i % 2 != 0]
                                
                                
                                
                                for i in range(len(time_objects)//2):
                                    
                                    start_timeoff = datetime.strptime("08:00:00","%H:%M:%S") if time_objects[i * 2]<=datetime.strptime("08:00:00","%H:%M:%S") else time_objects[i * 2]
                                    
                                    end_timeoff = datetime.strptime("17:00:00","%H:%M:%S") if time_objects[i * 2 + 1]>=datetime.strptime("17:00:00","%H:%M:%S") else time_objects[i * 2 + 1]
                                    
                                    timeoff_duration = end_timeoff - start_timeoff
                                    timeatten += math.floor(timeoff_duration.total_seconds() / 60)
                                
                                
                                
                                # morning_time=datetime.strptime("08:00:00","%H:%M:%S") if time_objects[0]<=datetime.strptime("08:00:00","%H:%M:%S") else time_objects[0]
                                # afternoon_time=datetime.strptime("17:00:00","%H:%M:%S") if time_objects[-1]>=datetime.strptime("17:00:00","%H:%M:%S") else time_objects[-1]
                                # time_difference = afternoon_time - morning_time
                                # working_time = math.floor(time_difference.total_seconds() / 60)
                                
                                
                                if time_objects[-1]<datetime.strptime("12:00:00","%H:%M:%S"):
                                    breaktime_duration-=60
                                
                                totaltimeatten=timeatten-breaktime_duration if timeatten-breaktime_duration >0 else 0
                                totaltimeoff=480-totaltimeatten
                                
                                timein=", ".join(timeins)
                                timeout=", ".join(timeouts)
                                sheet.append([day.day,timein,timeout,breaktimescan,totaltimeatten,totaltimeoff])
                                
                            else:
                                sheet.append([day.day])
                            
                            
                      
                                  
        
        current_date = datetime.now()
        current_month_name = current_date.strftime('%B')  
        workbook.save(f"staff_attendance{currect_year}({current_month_name}).xlsx")
        
    except Exception as e:
        print(f"Error occur is {e}")
        
def export_daily():
    try:
        workbook = openpyxl.Workbook()
        query="SELECT * FROM staff_list order by staff_name"
        cursor.execute(query)
        staffs=cursor.fetchall()
        sheet=workbook.active
        sheet["A1"]="STAFF ID"
        sheet["B1"]="STAFF NAME"
        sheet["C1"]="TIME IN"
        row_num = 2
        sheet.column_dimensions["A"].width=15
        sheet.column_dimensions["B"].width=40
        sheet.column_dimensions["C"].width=40
        current_date = datetime.now().date()
        green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")  # Green
        red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        send_status=0
        
        message="Today's Staff Attendance:\n"
        message+=f"Date: {current_date}\n"
        for staff in staffs:
            qry=f"SELECT  GROUP_CONCAT(time_checkin) as timecheckin FROM staff_attendance where DATE(date_checkin)=DATE('{current_date}') AND staff_id='{staff[3]}' AND (time_section='morning' or time_section='afternoon' or time_section='other') group by date_checkin"
            cursor.execute(qry)
            row=cursor.fetchone() 
            
            if row is None:
                message+=f"{staff[1]} : no attend today \n"
                send_status+=1
            
            attendance=row[0] if row is not None else "no attend today"
            sheet.append([staff[3],staff[1],attendance])
            
            cell = sheet[f"B{row_num}"]
            if attendance != "no attend today":
                cell.fill = green_fill  # Green if attended
            else:
                cell.fill = red_fill
            row_num += 1
        sendmessage=message
        # remind_whatapp(sendmessage,send_status)
        
        workbook.save(f"today staff attendance.xlsx")
        
        
    except Exception as e:
        print(f"Error occur is {e}")


def remind_whatapp(message,status=0):
    ACCESS_TOKEN = 'b4c1ff649fbc1173a2d03776a97860e22e77d87f4cf4235aba92b6b18ee54aa5'
    TO_PHONE_NUMBER = '60168813607'
    
    
    url = f"https://onsend.io/api/v1/send"
    headers = {
        'Accept': 'application/json',
        'Authorization': f'Bearer {ACCESS_TOKEN}',
        'Content-Type': 'application/json',
    }
    TO_PHONE_NUMBER='60129253398'
    
    data={
        'phone_number':TO_PHONE_NUMBER,
        'message':message,
    }
    if status>0:
        response = requests.post(url, headers=headers, json=data)
        if response.status_code == 200:
            print('Message sent successfully')
        else:
            print('Failed to send message:', response.json())
    
    
        
export_excel()
export_daily()
    
    